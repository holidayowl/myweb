"""将 data/ 下的 Excel 数据转为网站 localStorage 格式，不改动原始文件"""
import json, uuid, datetime, re, openpyxl

def uid():
    return str(uuid.uuid4())

def dtz(val):
    if isinstance(val, (datetime.datetime, datetime.date)):
        return val.strftime('%Y-%m-%d')
    return str(val) if val else ''

def to_num(v):
    if v is None: return None
    try: return float(v)
    except: return None

def clean_text(v):
    return str(v).strip() if v else ''

NOW = datetime.datetime.now().isoformat()

# ============================================================
# 1. 合同信息 (from 合同信息.xlsx)
# ============================================================
contracts = []
wc = openpyxl.load_workbook('data/合同信息.xlsx', data_only=True)
wsc = wc['Sheet1']
for r in range(2, wsc.max_row + 1):
    name = clean_text(wsc.cell(r, 1).value)
    cno  = clean_text(wsc.cell(r, 2).value)
    partner = clean_text(wsc.cell(r, 3).value)
    end_date = wsc.cell(r, 4).value
    amount_raw = wsc.cell(r, 5).value
    ctype  = clean_text(wsc.cell(r, 6).value)
    amt = to_num(amount_raw)
    notes = ''
    if amt is None:
        notes = f'金额: {amount_raw}'
        amt = 0
    if ctype:
        notes = (notes + ' | ' + ctype).strip(' | ')
    contracts.append({
        'id': uid(), 'contractName': name, 'contractNo': cno, 'partner': partner,
        'signDate': '', 'startDate': '', 'endDate': dtz(end_date) if end_date else '',
        'amount': amt, 'content': '', 'attachment': '', 'attachmentName': '',
        'status': 'active', 'notes': notes,
        'createdAt': NOW, 'updatedAt': NOW,
    })
wc.close()
print(f'合同: {len(contracts)} 条')

# ============================================================
# 2. 能耗信息 (from 2019-2026年能耗统计.xlsx)
#    同时提取用量和费用支出
# ============================================================
energy = []
TYPE_MAP = {'水费': 'water', '电费': 'electric', '燃气费': 'gas'}
UNIT_MAP = {'water': '吨', 'electric': '度', 'gas': '立方米'}

wb1 = openpyxl.load_workbook('data/2019-2026年能耗统计.xlsx', data_only=True)
for sheet_name, etype in TYPE_MAP.items():
    ws = wb1[sheet_name]
    years = []
    for c in range(2, 10):
        yr = clean_text(ws.cell(2, c).value).replace('年', '')
        if yr and yr.isdigit(): years.append(yr)

    months = ['01','02','03','04','05','06','07','08','09','10','11','12']

    for r in range(3, 15):
        m = months[r - 3]
        vol_base = 10  # volume columns start at col 10
        cost_base = 2  # cost columns start at col 2
        for i, year in enumerate(years):
            vol = to_num(ws.cell(r, vol_base + i).value)
            cost = to_num(ws.cell(r, cost_base + i).value)
            if vol is not None and vol > 0:
                energy.append({
                    'id': uid(), 'energyType': etype, 'period': f'{year}-{m}',
                    'value': vol, 'cost': cost if cost is not None else 0,
                    'unit': UNIT_MAP[etype], 'isAbnormal': False,
                    'notes': '', 'createdAt': NOW, 'updatedAt': NOW,
                })
wb1.close()

# 补充 2026年单月水电气明细 中有但上面缺失的记录
wb2 = openpyxl.load_workbook('data/2026年单月水电气明细.xlsx', data_only=True)
# Water
ws_w = wb2['水费']
for r in range(4, 16):
    m = str(r - 3).zfill(2)
    vol = to_num(ws_w.cell(r, 2).value)
    cost = to_num(ws_w.cell(r, 3).value)
    if vol and vol > 0:
        exist = [e for e in energy if e['period'] == f'2026-{m}' and e['energyType'] == 'water']
        if not exist and cost is not None:
            energy.append({
                'id': uid(), 'energyType': 'water', 'period': f'2026-{m}',
                'value': vol, 'cost': cost, 'unit': '吨', 'isAbnormal': False,
                'notes': '', 'createdAt': NOW, 'updatedAt': NOW,
            })
# Electric
ws_e = wb2['电费']
for r in range(4, 16):
    m = str(r - 3).zfill(2)
    vol = to_num(ws_e.cell(r, 7).value)
    cost = to_num(ws_e.cell(r, 4).value)
    if vol and vol > 0:
        exist = [e for e in energy if e['period'] == f'2026-{m}' and e['energyType'] == 'electric']
        if not exist and cost is not None:
            energy.append({
                'id': uid(), 'energyType': 'electric', 'period': f'2026-{m}',
                'value': vol, 'cost': cost, 'unit': '度', 'isAbnormal': False,
                'notes': '', 'createdAt': NOW, 'updatedAt': NOW,
            })
# Gas
ws_g = wb2['燃气费']
for r in range(2, 14):
    m = str(r - 1).zfill(2)
    vol = to_num(ws_g.cell(r, 6).value)
    cost = to_num(ws_g.cell(r, 14).value)  # 实际单价 column
    # cost is per-unit price, total cost = vol * price
    if vol and vol > 0 and cost is not None and cost > 0:
        total_cost = round(vol * cost, 2)
    else:
        total_cost = 0
    if vol and vol > 0:
        exist = [e for e in energy if e['period'] == f'2026-{m}' and e['energyType'] == 'gas']
        if not exist:
            energy.append({
                'id': uid(), 'energyType': 'gas', 'period': f'2026-{m}',
                'value': vol, 'cost': total_cost, 'unit': '立方米', 'isAbnormal': False,
                'notes': '', 'createdAt': NOW, 'updatedAt': NOW,
            })
wb2.close()
print(f'能耗: {len(energy)} 条')

# ============================================================
# 3. 物业运维支出 (from 3 property files)
# ============================================================
property_records = []
property_files = [
    'data/2024年物业支出明细 (自动保存的).xlsx',
    'data/2025年物业支出明细最新.xlsx',
    'data/2026年物业支出明细.xlsx',
]

# Classification keywords (order matters - first match wins)
CLASSIFY_RULES = [
    # 物业费（保安保洁礼仪接待）
    (['保洁', '清洁', '消杀', '除四害', '灭鼠', '灭蟑', '虫控', '防疫',
      '垃圾', '清运', '化粪池', '隔油池', '管道疏通', '清洗水池', '水箱清洗',
      '安保', '保安', '巡更', '对讲', '岗亭', '礼仪', '前台', '接待', '客服',
      '前台服务', '礼仪服务', '会务', '礼宾', '迎宾'], '物业费（保安保洁礼仪接待）'),
    # 房屋和设备维修费
    (['水管', '水暖', '水龙头', '阀门', '洁具', '灯具', '开关', '插座', '电线',
      '电缆', '灯管', '灯泡', '镇流器', '断路器', '继电器', '接触器', '熔断器',
      '管道', '接头', '弯头', '三通', '法兰', '密封', '垫片', '螺栓', '五金',
      '油漆', '涂料', '玻璃', '胶', '水泥', '砂石', '瓷砖', '地板', '地毯',
      '锁具', '铰链', '把手', '滑轨', '闭门器', '维修耗材', '材料',
      '水费', '电费', '水表', '电表', '给水', '排水', '供水', '供电', '用水',
      '水电维修', '水电改造', '水电安装', '水箱', '水质检测', '水质', '防水',
      '漏水', '渗水', '堵漏', '污水', '雨水', '冷凝水',
      '门窗', '幕墙', '防水', '屋面', '外墙', '渗漏', '裂缝', '沉降',
      '土建', '装修', '粉刷', '贴砖', '吊顶', '隔断', '围墙', '道路'], '房屋和设备维修费'),
    # 安全费
    (['安全', '消防', '灭火器', '消火栓', '喷淋', '烟感', '温感', '报警',
      '应急', '疏散', '防火', '防雷', '防爆', '防静电',
      '安全帽', '安全带', '安全网', '防护', '警示', '标识',
      '职业病', '体检', '安全培训', '安全评价', '安全检测',
      '电梯年检', '锅炉年检', '压力容器', '特种设备检验'], '安全费'),
    # 环境和设备维保费
    (['绿化', '花卉', '草坪', '树木', '修剪', '绿植', '盆景', '园林',
      '电梯', '空调', '锅炉', '冷却塔', '配电', '变压器', '高低压', '发电',
      '设备维保', '设备维护', '设备保养',
      'LED', '屏幕', '显示器', '监控', '摄像头', '门禁', '道闸',
      '排烟', '风机', '水泵', '电机', '压缩机', '控制柜', '变频', '弱电', '智能化',
      '机械', '立体车库', '停车设备', '擦窗机', '卷帘门', '伸缩门', '电动门',
      '环境检测', '环境监测', '环保', '排污', '噪声', '废气', '废水'], '环境和设备维保费'),
]

def classify_expense(item_name):
    """Classify expense item by keyword matching"""
    for keywords, etype in CLASSIFY_RULES:
        for kw in keywords:
            if kw in item_name:
                return etype
    return '其他'

def parse_property_file(fpath):
    """Parse a property expense file, return list of records"""
    records = []
    wb = openpyxl.load_workbook(fpath, data_only=True)
    year_sheets = [s for s in wb.sheetnames if s.isdigit() and 2020 <= int(s) <= 2026]

    for sname in year_sheets:
        ws = wb[sname]
        year = sname
        current_month = None

        for r in range(2, ws.max_row + 1):
            month_val = clean_text(ws.cell(r, 1).value)
            item_name = clean_text(ws.cell(r, 2).value)
            amount = to_num(ws.cell(r, 3).value)
            note = clean_text(ws.cell(r, 4).value)

            if not item_name:
                continue

            # Update current month if provided
            if month_val:
                month_map = {
                    '一月': '01', '二月': '02', '三月': '03', '四月': '04',
                    '五月': '05', '六月': '06', '七月': '07', '八月': '08',
                    '九月': '09', '十月': '10', '十一月': '11', '十二月': '12',
                }
                current_month = month_map.get(month_val, month_val)
                if not current_month or len(current_month) > 2:
                    # Try numeric
                    m = re.findall(r'\d+', month_val)
                    current_month = m[0].zfill(2) if m else None

            if amount is None or amount <= 0:
                # Check if note contains amount info
                if note:
                    note_amt = to_num(re.sub(r'[^0-9.]', '', note))
                    if note_amt:
                        amount = note_amt
                if amount is None or amount <= 0:
                    continue

            if current_month is None:
                continue

            expense_type = classify_expense(item_name)
            payment_method = '银行转账'

            records.append({
                'id': uid(),
                'item_name': item_name,
                'year': year,
                'month': current_month,
                'expenseDate': f'{year}-{current_month}-15',
                'expenseType': expense_type,
                'amount': amount,
                'paymentMethod': payment_method,
                'purpose': item_name,
                'isLargeExpense': amount >= 5000,
                'notes': note if note else '',
            })
    wb.close()
    return records

# Parse all 3 files
all_property = []
for fpath in property_files:
    try:
        recs = parse_property_file(fpath)
        all_property.extend(recs)
        print(f'  解析 {fpath}: {len(recs)} 条')
    except Exception as e:
        print(f'  错误 {fpath}: {e}')

# Deduplicate by (year, month, item_name, amount)
seen = set()
deduped = []
for r in all_property:
    key = (r['year'], r['month'], r['item_name'], r['amount'])
    if key not in seen:
        seen.add(key)
        deduped.append(r)

# Sort by date
deduped.sort(key=lambda x: (x['year'], x['month']))

# Convert to website schema
property_data = []
for r in deduped:
    settings = {'largeExpenseThreshold': 5000}
    property_data.append({
        'id': r['id'],
        'expenseType': r['expenseType'],
        'amount': r['amount'],
        'expenseDate': r['expenseDate'],
        'paymentMethod': r['paymentMethod'],
        'purpose': r['purpose'],
        'isLargeExpense': r['amount'] >= 5000,
        'notes': r['notes'],
        'createdAt': NOW,
        'updatedAt': NOW,
    })

print(f'物业运维支出: {len(deduped)} 条 (去重后)')

# ============================================================
# 4. 输出
# ============================================================
output = {
    'version': '1.0',
    'exportDate': datetime.datetime.now().isoformat(),
    'data': {
        'contracts': contracts,
        'energy': energy,
        'property': property_data,
        'equipment': [],
        'others': [],
        'others_schema': [
            {'key': 'workName', 'label': '工作名称', 'type': 'text', 'required': True, 'order': 1},
            {'key': 'recordDate', 'label': '记录时间', 'type': 'date', 'required': True, 'order': 2},
            {'key': 'category', 'label': '事项分类', 'type': 'select', 'required': True, 'options': ['台账记录', '对外对接', '会议纪要', '检查记录', '其他'], 'order': 3},
            {'key': 'attachment', 'label': '相关附件', 'type': 'file', 'required': False, 'order': 4},
            {'key': 'notes', 'label': '备注说明', 'type': 'textarea', 'required': False, 'order': 5},
        ],
        'settings': {
            'largeExpenseThreshold': 5000,
            'contractExpiryRemindDays': 30,
            'equipmentRecheckRemindDays': 7,
            'energyAbnormalPercent': 30,
        },
        'alert_dismissed': [],
        'sync_state': {'lastPush': None, 'lastPull': None},
    }
}

with open('data/import-data.json', 'w', encoding='utf-8') as f:
    json.dump(output, f, ensure_ascii=False, indent=2)

# Stats
from collections import Counter
etype_counts = Counter(r['expenseType'] for r in property_data)
print(f'物业分类: {dict(etype_counts)}')
eyear_counts = Counter(e['period'][:4] for e in energy)
print(f'能耗年份分布: {dict(sorted(eyear_counts.items()))}')
total_cost = sum(e.get('cost', 0) for e in energy)
print(f'能耗总费用: {total_cost:,.0f} yuan')
print(f'文件大小: {len(json.dumps(output, ensure_ascii=False)):,} bytes')
print('Done: data/import-data.json')

# ============================================================
# 5. 生成整理后的 Excel 文件
# ============================================================
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

def write_excel(filepath, sheets_data):
    wb = Workbook()
    wb.remove(wb.active)
    hfont = Font(bold=True, size=11, color='FFFFFF')
    hfill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    for sname, headers, rows in sheets_data:
        ws = wb.create_sheet(title=sname[:31])
        for c, h in enumerate(headers, 1):
            cell = ws.cell(1, c, h)
            cell.font = hfont; cell.fill = hfill
            cell.alignment = Alignment(horizontal='center')
        for r, row in enumerate(rows, 2):
            for c, val in enumerate(row, 1):
                ws.cell(r, c, val)
        for c in range(1, len(headers)+1):
            max_w = len(str(headers[c-1]))
            for ri in range(2, len(rows)+2):
                v = str(ws.cell(ri, c).value or '')
                if len(v) > max_w: max_w = len(v)
            ws.column_dimensions[get_column_letter(c)].width = min(max(max_w, 8), 40) + 4
    wb.save(filepath)
    print(f'  Saved: {filepath}')

# 合同
write_excel('data/合同目录汇总.xlsx', [('合同目录',
    ['合同名称','合同编号','合作方','签订日期','有效期起','有效期止','金额','状态','合同内容','备注'],
    [[c['contractName'],c['contractNo'],c['partner'],c['signDate'],c['startDate'],c['endDate'],
      c['amount'] if c['amount']>0 else '非固定金额',
      '有效' if c['status']=='active' else ('已终止' if c['status']=='terminated' else '已过期'),
      c['content'],c['notes']] for c in contracts]
)])

# 能耗
type_cn = {'water':'水','electric':'电','gas':'气'}
write_excel('data/能耗信息汇总.xlsx', [('能耗信息',
    ['能耗类型','统计周期','用量','单位','费用(元)','备注'],
    [[type_cn.get(e['energyType'],e['energyType']), e['period'], e['value'], e['unit'], e.get('cost',0), e['notes']]
     for e in sorted(energy, key=lambda x: (x['period'], x['energyType']))]
)])

# 物业运维（按年份分Sheet）
prop_by_year = {}
for r in sorted(property_data, key=lambda x: x['expenseDate']):
    y = r['expenseDate'][:4] if r['expenseDate'] else '未知'
    if y not in prop_by_year: prop_by_year[y] = []
    prop_by_year[y].append([r['expenseType'], r['amount'], r['expenseDate'], r['paymentMethod'], r['purpose'], r['notes']])

prop_sheets = [('全部汇总',
    ['支出项目','金额','日期','支付方式','用途','备注'],
    [[r['expenseType'], r['amount'], r['expenseDate'], r['paymentMethod'], r['purpose'], r['notes']]
     for r in sorted(property_data, key=lambda x: x['expenseDate'])]
)]
for y in sorted(prop_by_year.keys()):
    prop_sheets.append((f'{y}年',
        ['支出项目','金额','日期','支付方式','用途','备注'],
        prop_by_year[y]
    ))
write_excel('data/物业运维支出总表.xlsx', prop_sheets)
print('Excel files generated.')
